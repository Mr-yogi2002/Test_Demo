# excel_functions.py

# file where all excel read and write functions are there

from openpyxl import load_workbook

class Excel_Functions:
    # def __int__(self,excel_file_name,sheet_name):
    #     self.file = excel_file_name
    #     self.file = sheet_name
    #
    # # fetch the row_count
    # def row_count(self):
    #     workbook = load_workbook(self.file)
    #     sheet = workbook[self.sheet]
    #     return sheet.max_row
    #
    # # fetch the column_count
    # def column_count(self):
    #     workbook = load_workbook(self.file)
    #     sheet = workbook[self.sheet]
    #     return sheet.max_column
    #
    # #read the excel file
    # def read_data(self,row_number,column_number):
    #     workbook = load_workbook(self.file)
    #     sheet = workbook[self.sheet]
    #     return sheet.cell(row=row_number,column=column_number).value
    #
    # # read data into excel file
    # def write_data(self,row_number,column_number,data):
    #     workbook = load_workbook(self.file)
    #     sheet = workbook[self.sheet]
    #     sheet.cell(row=row_number,column=column_number).value=data
    #     workbook.save(self.file)

    def __init__(self, excel_file_name, sheet_name):
        self.file = excel_file_name
        self.sheet = sheet_name

    def Row_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_row)

    def Column_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_column)

    def Read_Data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.cell(row=row_number, column=column_number).value)

    def Write_Data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        data = sheet.cell(row=row_number, column=column_number).value
        workbook.save(self.file)
